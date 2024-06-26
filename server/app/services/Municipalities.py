# Service for fetching information about dangerous buildings from municipal websites

import json
from selenium import webdriver
import requests
import defusedxml.ElementTree as ET
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import date, datetime
from utiles.secure_openpyxl import read_xlsx_data

file_path = '236.xlsx'
data_path = f'/home/daniel/projects/{file_path}'

def fetch_and_save_file(url, output_file):
    try:
        # Fetch the XML content from the URL
        response = requests.get(url)
        response.raise_for_status()  # Raise an HTTPError if the
        
        
        with open(output_file, 'wb') as f:
            f.write(response.content)

        print(f"XML data has been saved to {output_file}")

    except requests.exceptions.RequestException as e:
        print(f"Error fetching the XML data: {e}")

    except ET.ParseError as e:
        print(f"Error parsing the XML data: {e}")
    
# Reading the saved XLSX file to verify
# def unsecure_read_xlsx_data(file_path):
#     try:
#         # Load the workbook and select the active worksheet
#         workbook = load_workbook(file_path)
#         sheet = workbook.active
        
#         # Iterate through rows and columns to read data
#         data = []
#         for row in sheet.iter_rows(values_only=True):
#             if any(cell is not None for cell in row):
#                 data.append(row)
        
#         return data
    
#     except Exception as e:
#         print(f"Error reading the XLSX data: {e}")
#         return None
    

def split_buildings_by_address(data):
    '''
    data[1] = ("אבא ברדיצ'ב", "12 א'+ב'", 2, datetime.datetime(2023, 11, 26, 0, 0), None, None)
    
    data[2] = (None, 16, 1, datetime.datetime(2017, 11, 23, 0, 0), None, None)
    '''
    current_address = data[1][0]
    result = {current_address:[{"building_num": data[1][1],
                               "risk_group":  data[1][2],
                               "announcement_date": data[1][3].strftime('%Y-%m-%d'),
                               "announcement_status": data[1][4],
                               "notes": data[1][5]}]
              }
    
    # For each row (building) save building details under the corrent address
    prev_row = result[current_address][0]
    for row in data[2:]:
        row_address = row[0]
        if row_address:
            current_address = row[0]
        
        if 'שינדלר' in current_address and not row[1]:
            pass
        building_dict =  create_building_dict(row, prev_row)
        # result[current_address].append(building_dict)
        result.setdefault(current_address, []).append(building_dict)
        prev_row = building_dict
    return result

def create_building_dict(data, prev_row):
    result = {}
    try:
        
        result = {"building_num": data[1] if data[1] else prev_row['building_num'],
                    "risk_group":  data[2] if data[2] else prev_row['risk_group'],
                    "announcement_date": data[3].strftime('%Y-%m-%d') if data[3] and isinstance(data[3], (date, datetime)) else prev_row['announcement_date'],
                    "announcement_status": data[4],
                    "notes": data[5]}
    
    except AttributeError as e:
        print(f"{e}")
    return result

# Function to set up the Selenium WebDriver with options
def setup_browser():
    options = webdriver.FirefoxOptions()
    # Optional: Add other preferences to make the browser appear more human
    options.set_preference("dom.webdriver.enabled", False)
    options.set_preference('useAutomationExtension', False)
    options.set_preference("general.platform.override", "")
    
    # browser = webdriver.Firefox(options=options)
    
    return webdriver.Firefox(options=options)

def process_page(browser):
    try:
        # Open the target URL
        browser.get('https://www.bat-yam.muni.il/info.php?id=10389') 
        
        # Wait for the specific <ul> element to be present
        element_list = WebDriverWait(browser, 15).until(
            EC.presence_of_element_located((By.XPATH, '//div[@class="legoText_10389_cnt_body_heb"]'))
        )
        
        # Locate the <a> element within the specific <div> that has an href attribute
        link_element = element_list.find_element(By.XPATH, './/a[@href]')
        href_value = link_element.get_attribute('href')
        
        return href_value

        
    except TimeoutException:
        print("Timeout occurred while waiting for the element.")
    except NoSuchElementException:
        print("The element was not found in the DOM.")
    except Exception as e:
        print(f"An error occurred: {e}")

def get_xml_url_by_scraping():
    browser = setup_browser()
    xml_url = []
    try:
        xml_url = process_page(browser)
        pass 
   
    finally:
        browser.quit()
        pass
    
    return xml_url
    
def main():
    xml_url = get_xml_url_by_scraping()
    
    output_file_path = "bat-yam-dangerous-builds.xlsx"
    fetch_and_save_file(xml_url, output_file_path)
    
    # Reading the data from the output file
    data = read_xlsx_data(output_file_path)
    data_by_address = split_buildings_by_address(data)
    # print(data_by_address)
    import itertools
    import pprint
    first_4_elements = dict(itertools.islice(data_by_address.items(), 4))
    print(pprint.pprint(first_4_elements))
    
    
    

if __name__ == "__main__":
    main()