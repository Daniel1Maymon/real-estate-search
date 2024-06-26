from defusedxml.ElementTree import parse
from defusedxml.ElementTree import fromstring

import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import ExcelReader, load_workbook

'''
We override the parse and fromstring methods used by openpyxl to use the secure versions provided by defusedxml.ElementTree.
'''
# Patch openpyxl to use defusedxml
ExcelReader.parse = staticmethod(parse)
ExcelReader.fromstring = staticmethod(fromstring)

def read_xlsx_data(file_path):
    try:
        # Load the workbook and select the active worksheet
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Iterate through rows and columns to read data
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)

        return data

    except Exception as e:
        print(f"Error reading the XLSX data: {e}")
        return None